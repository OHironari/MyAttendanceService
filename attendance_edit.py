from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from datetime import datetime,date,time
from get_excel_info_from_s3 import load_file_from_s3
from dotenv import load_dotenv
import os,io,re
import pandas as pd

class workrecord:
    def __init__(self, work_date=None, start_time=None, end_time=None):
        if work_date:
            self.work_date = datetime.strptime(work_date, "%Y-%m-%d")
        else:
            self.work_date = datetime.combine(date.today(), time.min)

        if start_time:
            self.start_time = datetime.strptime(start_time, "%H:%M").time()
        else:
            self.start_time = time(9, 0)  # 9:00

        if end_time:
            self.end_time = datetime.strptime(end_time, "%H:%M").time()
        else:
            self.end_time = time(17, 30)  # 17:30
    def __str__(self):
        return f"{self.work_date} | {self.start_time} - {self.end_time} 時間"



def main(work_record,wb):
    # # WorkBookの読み込み
    # wb = load_workbook("../attendance_202506.xlsx",data_only=True)
    df = pd.read_excel(io.BytesIO(wb))

if __name__ == "__main__":
    # .envを読み込む
    load_dotenv()
    bucket_name=os.getenv("bucket_name")

    input_starttime=input('starttime(hh:mm):')
    input_endtime=input('endtime(hh:mm):')
    input_date=input('date(yyyymmdd):')

    work_record = workrecord(
        work_date=input_date,
        start_time=input_starttime,
        end_time=input_endtime
    )
    wb = load_file_from_s3(bucket_name)
    main(work_record,wb)

