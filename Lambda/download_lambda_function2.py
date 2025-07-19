import boto3
import openpyxl
from openpyxl.styles import Alignment, PatternFill
import io
import os
import logging
from datetime import datetime
import json
import urllib.parse

logger = logging.getLogger()
logger.setLevel(logging.INFO)

#重要度は低いので、lambdaの環境変数から取得
bucket_name=os.getenv("bucket_name")
secrets_manager_arn=os.getenv("secrets_manager_arn")
template_path=os.getenv("template_filepath")

s3 = boto3.client("s3")

def lambda_handler(event, context):
    try:
        headers = {k.lower(): v for k, v in event.get("headers", {}).items()}
        content_type = headers.get("content-type", "")
        if "application/json" in content_type:
            body = json.loads(event["body"])
        else:
            parsed = urllib.parse.parse_qs(event.get("body") or "")
            body = {k: v[0] for k, v in parsed.items()}

        # Check Credential
        check = check_credential(body["id_token"])
        if check.get("status") == "false":
            logger.info("token expired")
            return response(401, {'error': 'token expired', 'records': []})
        sub = check.get("sub")

        body["sub"] = str(sub)


        # get assume role arn from secrets manager
        secrets_dict = get_role_arn_s3()
        role_arn = secrets_dict['s3_access_role_arn']
        UserName = secrets_dict['UserName']
        logger.info(f"role_arn:{role_arn}")
        if role_arn is None or isinstance(role_arn, dict):
            return response(500, {'error': 'Failed to get role arn', 'detail': role_arn})

        # assume role
        session=get_role(role_arn)
        logger.info(f"session:{session}")
        if isinstance(session,dict) and "error" in session:
            return response(500,session)

        #MyAttendanceFunctionにwork_dateの見渡すことで一覧を取得する
        result = get_attendance_list(body,session)

        # resutlの結果をテンプレートファイルに記述 署名付きURLを発行
        urls = write_list_to_excel(UserName,result,session)

        #return response(200, result)
        return response(200, {"url": urls})

    except Exception as e:
        logger.info(f"lambda_hander_error:{str(e)}")
        return response(500, {'error': str(e)})

def get_role_arn_s3():
    try:
        # Secrets Manager クライアント作成
        sct_mgr = boto3.client(service_name='secretsmanager', region_name='us-east-1')
        
        # Secret 取得
        secret_value = sct_mgr.get_secret_value(SecretId=secrets_manager_arn)
        
        # JSON 文字列を辞書に変換
        secret_dict = json.loads(secret_value['SecretString'])

        # s3_access_role_arn を返す
        return secret_dict

    except Exception as e:
        return {"error": str(e)}


def get_role(role_arn):
    try:
        # STS クライアント作成
        sts = boto3.client('sts', region_name='ap-northeast-1')
        # AssumeRole 実行
        response = sts.assume_role(
            RoleArn=role_arn,
            RoleSessionName="s3access",
            # DurationSeconds=900  # ← 必要なら有効化
        )
        
        # 新しいセッション作成
        session = boto3.Session(
            aws_access_key_id=response['Credentials']['AccessKeyId'],
            aws_secret_access_key=response['Credentials']['SecretAccessKey'],
            aws_session_token=response['Credentials']['SessionToken'],
            region_name='ap-northeast-1'
        )
        
        return session

    except Exception as e:
        return {"error": str(e)}

def response(status_code, body_dict):
    logger.info(f"status_code:{status_code},body_dict:{body_dict}")
    return {
        "statusCode": status_code,
        "headers": {
            "Content-Type": "application/json; charset=utf-8"
        },
        "body": json.dumps(body_dict, ensure_ascii=False)
    }

def get_attendance_list(body,session):
    try:
        lambda_client = session.client("lambda")

        if body.get("work_date"):
            work_date_obj = datetime.strptime(body["work_date"], "%Y-%m-%d")
            work_date_str = work_date_obj.strftime("%Y-%m-%d")
        else:
            work_date_str = ""

        # payloadは相手のLambdaがAPI Gateway経由でのアクセスを想定していたため、この形にしてリクエストを送る
        payload = {
            "headers": {
                "Content-Type": "application/json"
            },
            "body": json.dumps({
            "user_id": "iamuser",
            "work_date": work_date_str,
            'day_of_the_week': "",
            'work_style': "",
            'start_time': "",
            'end_time': "",
            'break_time': "",
            'work_time': "",
            'note': "",
            'sub': body.get("sub"),
            'id_token': body.get("id_token"),
            'style': "readonly"
            })
        }

        response = lambda_client.invoke(
            FunctionName="MyAttendanceFunction",
            InvocationType="RequestResponse",
            Payload=json.dumps(payload)
        )
        
        response_payload = json.loads(response["Payload"].read())

        # 呼び出し先 Lambda の body 部分をパース
        nested_body = response_payload.get("body")
        if nested_body:
            parsed_body = json.loads(nested_body)
        else:
            parsed_body = {}

        # records を直接取り出す
        records = parsed_body.get("records", [])

        return {
            "message": parsed_body.get("message", "Success"),
            "records": records
        }

    except Exception as e:
        logger.info(f"error:{str(e)}")
        raise
        

def write_list_to_excel(username,result,session):
    try:
        # テンプレート取得
        s3_client = session.client('s3')
        response = s3_client.get_object(Bucket=bucket_name, Key=template_path)
        template_data = response["Body"].read()
        in_mem_file = io.BytesIO(template_data)

        # Excel 読み込み
        workbook = openpyxl.load_workbook(in_mem_file)

        # 日付の定義
        date_obj = datetime.strptime(result["records"][0]["work_date"], "%Y-%m-%d")
        year = date_obj.strftime("%Y")
        month = date_obj.strftime("%m")
        day = date_obj.strftime("%d")

        #===============
        # 末締め作業実績深夜つき シート編集
        #===============
        sheet = workbook['末締め作業実績深夜つき']
        start_row = 8  # 書き込み開始行（テンプレートに合わせて適宜変更）
        sheet["H6"] = datetime.strptime(result["records"][0]["work_date"].replace("-", "/"), "%Y/%m/%d")
        sheet["T5"] = username

        #一覧部分
        for idx, record in enumerate(result["records"]):
            row = start_row + idx

            # 例として以下のカラムに書く
            #sheet[f"A{row}"] = record.get("work_date", "")
            #sheet[f"Z{row}"] = record.get("day_of_the_week", "")
            sheet[f"V{row}"] = "I区" if record.get("work_style", "") == "出勤" else None
            sheet[f"N{row}"] = datetime.strptime(record.get("start_time", ""),"%H:%M").time() if record.get("start_time", "") else None
            sheet[f"P{row}"] = datetime.strptime(record.get("end_time", ""),"%H:%M").time() if record.get("end_time", "") else None
            #sheet[f"AD{row}"] = datetime.strptime(record.get("break_time", ""),"%H:%M").time() if record.get("break_time", "") else None
            #sheet[f"AE{row}"] = datetime.strptime(record.get("work_time", ""),"%H:%M").time() if record.get("work_time", "") else None
            sheet[f"S{row}"] = record.get("note", "")
            if record.get("work_style", "") == "休み":
                cell = sheet[f"L{row}"]
                cell.fill = PatternFill(fill_type='solid', start_color='FF6900', end_color='FF6900')

        # 書き込み後、ファイルをバイトデータに保存
        out_mem_file = io.BytesIO()
        workbook.save(out_mem_file)
        out_mem_file.seek(0)  # 先頭に戻す

        #===============
        # 近地出張旅費 シート編集
        #===============
        sheet = workbook['近地出張旅費']

        title_str=f"近地等出張旅費計算書({year}年{month}月)"
        request_date = f"依頼日　　　{year}年　　 {month}月{day}日"

        sheet["B3"] = title_str
        sheet["J7"] = username
        sheet["I5"] = request_date
        cell = sheet["B3"]
        cell.alignment = Alignment(horizontal="left", vertical="center")

        # 書き込み後、ファイルをバイトデータに保存
        out_mem_file = io.BytesIO()
        workbook.save(out_mem_file)
        out_mem_file.seek(0)  # 先頭に戻す

        #===============
        # 業務完了報告書 シート編集
        #===============
        sheet = workbook['業務完了報告書']
        logger.info(f"sheet:{sheet}")

        # start date オブジェクトに変換
        start_date = datetime.strptime(result["records"][0]["work_date"], "%Y-%m-%d")
        start_year  = start_date.strftime("%Y")
        start_month = start_date.strftime("%m")
        start_day   = start_date.strftime("%d")

        # end date オブジェクトに変換
        end_date = datetime.strptime(result["records"][-1]["work_date"], "%Y-%m-%d")
        end_year  = end_date.strftime("%Y")
        end_month = end_date.strftime("%m")
        end_day   = end_date.strftime("%d")

        term_str = f"{start_year}年{start_month}月{start_day}日　〜　{end_year}年{end_month}月{end_day}日"

        logger.info(f"term_str{term_str}")

        sheet["N8"] = f"報告者：{username}"
        sheet["N4"] = end_year
        sheet["Q4"] = end_month
        sheet["S4"] = end_day
        sheet["C14"] = term_str
        sheet["F50"] = datetime.strptime(result["records"][-1]["work_date"].replace("-", "/"), "%Y/%m/%d")

        # 書き込み後、ファイルをバイトデータに保存
        out_mem_file = io.BytesIO()
        workbook.save(out_mem_file)
        out_mem_file.seek(0)  # 先頭に戻す

        formatted = date_obj.strftime("%Y%m")
        s3_client.put_object(
            Bucket=bucket_name,
            Key=f"attendance/files/downloads/作業実績表_{username}_{formatted}.xlsx",
            Body=out_mem_file.getvalue()
        )

        # 署名付きURLの発行
        url = s3_client.generate_presigned_url(
                "get_object",
                Params={"Bucket": bucket_name, "Key": f"attendance/files/downloads/作業実績表_{username}_{formatted}.xlsx"},
                ExpiresIn=3600
            )
        
        return url

    except Exception as e:
        logger.info(f"error:{str(e)}")
        raise

def check_credential(id_token):
    try:
        payload = {"idtoken": id_token, "invokefunction": "Download Function2"}
        logger.info(f"payload:{payload}")
        lambda_client = boto3.client("lambda")
        response = lambda_client.invoke(
            FunctionName='CheckCredentialFunction',
            InvocationType='RequestResponse',
            Payload=json.dumps(payload)
        )
        payload_str = response["Payload"].read().decode("utf-8")
        payload_dict = json.loads(payload_str)
        body_str = payload_dict.get("body", "{}")
        body_dict = json.loads(body_str)
        status_code = payload_dict.get("statusCode", 500)
        if status_code == 200 and body_dict.get("status") == "online":
            return {"status": "true", "sub": body_dict.get("sub"), "email": body_dict.get("email")}
        else:
            logger.info(f"Invalid credential: {body_dict.get('error', 'Unknown error')}")
            return {"status": "false"}
    except Exception as e:
        logger.info(f"check_credential error: {str(e)}")
        return {"status": "false"}

# curl https://app.itononari.xyz/download -XPOST -H "Content-Type: application/json" -d '{"work_date":"2025-07-04", "start_time":"09:00", "end_time":"18:15"}'
